# ============================================================
#  MOBILISIS - klijent za Mobilisis Fleet (GPS pozicije vozila)
#
#  API: https://fleet2.mobilisis.hr/geocodeAndZoneAPI/api/v1
#    POST /positions/getSessionKey  -> token (traje 24h, produzuje se koristenjem)
#    GET  /positions/devices        -> lista uredaja {Id, Name(=registracija)}
#    GET  /positions/position       -> trenutne pozicije svih uredaja
#  (putanje potvrdene prema swaggeru: .../geocodeAndZoneAPI/swagger/docs/v1)
#
#  - Token se kesira u memoriji (modul-level). Ako bilo koji poziv vrati 401,
#    napravi se novi login i poziv se ponovi JEDNOM.
#  - Header: Authorization: Bearer <token>
#  - Timeout 15s, retry 2x (mrezne greske / 5xx).
#
#  Mapiranje GB <-> REG: iz Excela "GARAŽNI BROJEVI.xlsx" na istom SharePoint
#  siteu/biblioteci kao Racuni_terena (preko graph_client). Kes 24h.
#
#  Kredencijali iz okoline:
#    MOBILISIS_USER, MOBILISIS_PASS  (obavezno)
#    MOBILISIS_APP_ID                (opcionalno; swagger loginData.applicationId)
# ============================================================

import io
import os
import re
import threading
import time
from datetime import datetime, timedelta, timezone

import requests

import graph_client
import monitoring

# ---- Konfiguracija ----
BASE_URL = "https://fleet2.mobilisis.hr/geocodeAndZoneAPI/api/v1"
POSITIONS_PATH = "/positions/position"   # GET trenutne pozicije (Positions_GetPositions)

_USER = os.getenv("MOBILISIS_USER", "").strip()
_PASS = os.getenv("MOBILISIS_PASS", "").strip()
_APP_ID = os.getenv("MOBILISIS_APP_ID", "").strip()  # opcionalno

_TIMEOUT = 15
_RETRIES = 2

MAP_FILE = "GARAŽNI BROJEVI.xlsx"
MAP_TTL = 24 * 3600   # osvjezi mapiranje ako je starije od 24h

# ---- Kes tokena ----
_token = None
_token_lock = threading.Lock()

# ---- Kes mapiranja GB<->REG ----
_map = {"ts": 0.0, "gb2reg": {}, "reg2gb": {}, "regs": []}
_map_lock = threading.Lock()


class MobilisisError(Exception):
    pass


def is_configured():
    return bool(_USER and _PASS)


def _log(msg):
    print(f"[mobilisis] {msg}", flush=True)


# ==================== POMOCNE ====================

def _txt(v):
    return "" if v is None else str(v).strip()


def norm_reg(s):
    """Registracija: ukloni razmake i crtice, uppercase.
    'ZG 5267-KM' -> 'ZG5267KM'."""
    return re.sub(r"[\s\-]+", "", _txt(s)).upper()


def norm_gb(s):
    """GB kao string bez vodecih nula. Excel float '12.0' -> '12', '007' -> '7'."""
    t = _txt(s)
    if not t:
        return ""
    m = re.fullmatch(r"(\d+)(?:\.0+)?", t)
    if m:
        t = m.group(1)
    return t.lstrip("0") or "0"


def _num_dot(v):
    """Broj (int/float) ili None. String sa zarezom -> tocka."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = _txt(v).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def ignition_on(v):
    """Je li motor upaljen (ignitionState raznih oblika) -> bool."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    return _txt(v).lower() in ("1", "true", "on", "yes", "upaljen",
                               "ignition_on", "ignitionon")


def _as_list(data, *keys):
    """Izvuci listu iz odgovora koji moze biti lista ili omotan u dict."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k in keys + ("value", "data", "Data", "result", "Result", "items"):
            v = data.get(k)
            if isinstance(v, list):
                return v
    return []


def parse_utc(s):
    """Parsiraj Mobilisis dateTime (UTC!) u aware datetime u UTC, ili None."""
    t = _txt(s)
    if not t:
        return None
    dt = None
    try:
        dt = datetime.fromisoformat(t.replace("Z", "+00:00"))
    except ValueError:
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(t, fmt)
                break
            except ValueError:
                continue
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


# ==================== HTTP ====================

def _http(method, url, **kw):
    """requests s timeoutom i retryjem (2x) na mrezne greske / 5xx."""
    kw.setdefault("timeout", _TIMEOUT)
    last = None
    for attempt in range(_RETRIES + 1):
        try:
            resp = requests.request(method, url, **kw)
            if resp.status_code >= 500 and attempt < _RETRIES:
                time.sleep(1 + attempt)
                continue
            return resp
        except requests.RequestException as e:
            last = e
            if attempt < _RETRIES:
                time.sleep(1 + attempt)
                continue
    raise MobilisisError(f"Mrežna greška prema Mobilisisu: {last}")


# ==================== AUTENTIKACIJA ====================

def _extract_token(resp):
    """Izvuci token iz odgovora getSessionKey (string ili dict s poljem)."""
    try:
        data = resp.json()
    except ValueError:
        return _txt(resp.text).strip('"') or None
    if isinstance(data, str):
        return data.strip().strip('"') or None
    if isinstance(data, dict):
        for k in ("sessionKey", "SessionKey", "sessionkey", "token", "Token",
                  "access_token", "accessToken", "key", "Key"):
            v = data.get(k)
            if v:
                return str(v)
    return None


def _login_request(include_app_id):
    """Jedan POST getSessionKey. include_app_id -> doda applicationId iz env-a."""
    body = {"username": _USER, "password": _PASS}
    if include_app_id and _APP_ID.isdigit():
        body["applicationId"] = int(_APP_ID)
    return _http("POST", f"{BASE_URL}/positions/getSessionKey", json=body)


def _needs_app_id(resp):
    """True ako login odgovor sugerira da treba applicationId: 401 ili poruka o
    nedostajucem/obaveznom polju (applicationId)."""
    if resp.status_code == 401:
        return True
    txt = _txt(resp.text).lower()
    return ("applicationid" in txt or "application id" in txt
            or "missing" in txt or "required" in txt
            or "nedostaje" in txt or "obavezno" in txt or "obavezan" in txt)


def login():
    """POST /positions/getSessionKey -> token (24h). Salje SAMO user/pass;
    tek ako login vrati 401 ili gresku o nedostajucem polju, a MOBILISIS_APP_ID
    je postavljen, ponavlja s applicationId. Sprema token u modul-kes."""
    if not is_configured():
        raise MobilisisError(
            "Mobilisis kredencijali nisu postavljeni (MOBILISIS_USER/PASS).")
    global _token
    _log("login: getSessionKey…")
    resp = _login_request(include_app_id=False)

    if resp.status_code != 200 and _APP_ID.isdigit() and _needs_app_id(resp):
        _log(f"login (user/pass) -> {resp.status_code}; ponavljam s MOBILISIS_APP_ID")
        resp = _login_request(include_app_id=True)

    if resp.status_code != 200:
        raise MobilisisError(
            f"Login nije uspio ({resp.status_code}): {_txt(resp.text)[:200]}")
    token = _extract_token(resp)
    if not token:
        raise MobilisisError("Login: token nije pronađen u odgovoru.")
    with _token_lock:
        _token = token
    _log("login: OK")
    return token


def _auth_request(method, path, params=None, json_body=None, _retried=False):
    """Autenticirani poziv (GET s params ili POST s json_body). Na 401 -> relogin
    i JEDAN ponovni pokušaj."""
    with _token_lock:
        tok = _token
    if not tok:
        tok = login()
    kw = {"headers": {"Authorization": f"Bearer {tok}"}}
    if params is not None:
        kw["params"] = params
    if json_body is not None:
        kw["json"] = json_body
    resp = _http(method, f"{BASE_URL}{path}", **kw)
    if resp.status_code == 401 and not _retried:
        _log("401 -> relogin i ponovni pokušaj")
        login()
        return _auth_request(method, path, params, json_body, _retried=True)
    if resp.status_code != 200:
        raise MobilisisError(
            f"{method} {path} -> {resp.status_code}: {_txt(resp.text)[:200]}")
    return resp


# ==================== API POZIVI ====================

def get_devices():
    """GET /positions/devices -> [{'Id':.., 'Name':..}] (Name = registracija)."""
    resp = _auth_request("GET", "/positions/devices")
    out = []
    for d in _as_list(resp.json(), "Devices", "devices"):
        if not isinstance(d, dict):
            continue
        did = d.get("Id", d.get("id"))
        if did is not None:
            out.append({"Id": did, "Name": _txt(d.get("Name", d.get("name")))})
    return out


def _parse_position(p):
    coord = p.get("coordinate") or p.get("Coordinate") or {}
    if not isinstance(coord, dict):
        coord = {}
    return {
        "deviceId": p.get("deviceId", p.get("DeviceId", p.get("Id"))),
        "lat": _num_dot(coord.get("Latitude", coord.get("latitude"))),
        "lon": _num_dot(coord.get("Longitude", coord.get("longitude"))),
        "heading": _num_dot(p.get("heading", p.get("Heading"))),
        "speed": _num_dot(p.get("speed", p.get("Speed"))),
        "ignition": p.get("ignitionState", p.get("IgnitionState")),
        "dateTime": p.get("dateTime", p.get("DateTime")),
        "odometer": _num_dot(p.get("odometer", p.get("Odometer"))),
    }


def get_positions(device_ids=None):
    """GET /positions/position -> lista trenutnih pozicija (parsirano). Ako je
    device_ids zadan, filtrira klijentski po deviceId."""
    resp = _auth_request("GET", POSITIONS_PATH)
    parsed = [_parse_position(p) for p in _as_list(resp.json(), "Positions", "positions")
              if isinstance(p, dict)]
    if device_ids is not None:
        want = {str(x) for x in device_ids}
        parsed = [p for p in parsed if str(p.get("deviceId")) in want]
    return parsed


# ==================== POVIJEST (ruta vozila) ====================

def _iso_utc(dt):
    """datetime -> 'YYYY-MM-DDTHH:MM:SSZ' (UTC) za Mobilisis from/to."""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _parse_history_point(p):
    """Jedna točka povijesti -> {lat, lon, vrijeme, brzina, smjer} ili None.
    Mobilisis History može vratiti isti oblik kao trenutne pozicije
    (coordinate.Latitude…) ili plosnati (Lat/Lon/Time) — pokrivamo oba."""
    if not isinstance(p, dict):
        return None
    coord = p.get("coordinate") or p.get("Coordinate") or {}
    if not isinstance(coord, dict):
        coord = {}
    lat = _num_dot(coord.get("Latitude", coord.get("latitude")))
    lon = _num_dot(coord.get("Longitude", coord.get("longitude")))
    if lat is None:
        lat = _num_dot(p.get("Lat", p.get("lat", p.get("Latitude", p.get("latitude")))))
    if lon is None:
        lon = _num_dot(p.get("Lon", p.get("lon", p.get("Longitude", p.get("longitude")))))
    if lat is None or lon is None:
        return None
    dt = parse_utc(p.get("dateTime", p.get("DateTime", p.get("Time", p.get("time")))))
    return {
        "lat": lat,
        "lon": lon,
        "vrijeme": dt.astimezone(timezone.utc).isoformat() if dt else None,
        "brzina": _num_dot(p.get("speed", p.get("Speed"))),
        "smjer": _num_dot(p.get("heading", p.get("Heading", p.get("course", p.get("Course"))))),
    }


def get_history(device_id, frm, to):
    """POST /positions/history -> lista točaka {lat, lon, vrijeme, brzina, smjer}
    za uređaj u periodu [frm, to] (aware datetime). Sortirano po vremenu; točke
    bez koordinata se preskaču."""
    body = {"id": int(device_id), "from": _iso_utc(frm), "to": _iso_utc(to)}
    resp = _auth_request("POST", "/positions/history", json_body=body)
    data = resp.json()
    raw = _as_list(data, "Positions", "positions", "History", "history",
                   "Items", "items", "Route", "route", "Points", "points")
    if not raw and isinstance(data, list):
        raw = data
    tocke = [t for t in (_parse_history_point(p) for p in raw) if t]
    tocke.sort(key=lambda t: t.get("vrijeme") or "")
    return tocke


def id_za_reg(reg):
    """Registracija -> Mobilisis device Id (ili None). Usporedba normalizirana."""
    target = norm_reg(reg)
    if not target:
        return None
    for d in get_devices():
        if norm_reg(d.get("Name", "")) == target:
            return d.get("Id")
    return None


def putanja_za_reg(reg, sati=6):
    """Ruta vozila (po registraciji) za zadnjih 'sati' sati.
    Vraća {reg, id, od, do, tocke:[…]} ili baca MobilisisError. 'tocke' su
    prazne ako Mobilisis nema zapisa za taj period."""
    did = id_za_reg(reg)
    if did is None:
        raise MobilisisError(f"Nepoznata registracija: {reg}")
    to = datetime.now(timezone.utc)
    frm = to - timedelta(hours=float(sati))
    tocke = get_history(did, frm, to)
    return {"reg": reg, "id": did, "od": _iso_utc(frm), "do": _iso_utc(to),
            "tocke": tocke}


# ==================== MAPIRANJE GB <-> REG (Excel) ====================

def _download_map_file():
    """Skini 'GARAŽNI BROJEVI.xlsx' (prvo iz BRAVEL foldera, pa search po
    biblioteci ako nije tamo)."""
    try:
        return graph_client.download_named(MAP_FILE, folder=graph_client.FOLDER)
    except graph_client.GraphError as e:
        if e.status_code == 404:
            _log(f"'{MAP_FILE}' nije u {graph_client.FOLDER} -> tražim po biblioteci")
            item = graph_client.find_item_by_name(MAP_FILE)
            if item:
                return graph_client.download_item(item["id"])
        raise


def _find_header(ws, max_scan=15):
    """Nadji red zaglavlja s kolonama 'GB' i 'REG OZNAKA'. Vrati
    (header_row, gb_col_idx, reg_col_idx) ili (None, None, None)."""
    r = 0
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        r += 1
        gb_i = reg_i = None
        for i, c in enumerate(row):
            cu = _txt(c).upper()
            if cu == "GB" and gb_i is None:
                gb_i = i
            elif cu == "REG OZNAKA" and reg_i is None:
                reg_i = i
        if gb_i is not None and reg_i is not None:
            return r, gb_i, reg_i
    return None, None, None


def _build_map():
    """Ucitaj mapiranje iz Excela: nadji worksheet s kolonama GB i REG OZNAKA."""
    from openpyxl import load_workbook
    content = _download_map_file()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    gb2reg, reg2gb, regs = {}, {}, []
    try:
        for ws in wb.worksheets:
            hr, gb_i, reg_i = _find_header(ws)
            if hr is None:
                continue
            for row in ws.iter_rows(min_row=hr + 1, values_only=True):
                gb = norm_gb(row[gb_i]) if gb_i < len(row) else ""
                reg_disp = _txt(row[reg_i]) if reg_i < len(row) else ""
                reg = norm_reg(reg_disp)
                if reg_disp:
                    regs.append(reg_disp)
                if gb and reg:
                    gb2reg[gb] = reg
                    reg2gb[reg] = gb
            break  # nasli smo pravi worksheet -> dosta
    finally:
        wb.close()
    if not gb2reg:
        raise MobilisisError(
            f"U '{MAP_FILE}' nisam našao tablicu s kolonama 'GB' i 'REG OZNAKA'.")
    _log(f"mapiranje ucitano: {len(gb2reg)} vozila")
    return gb2reg, reg2gb, regs


def _get_map(force=False):
    now = time.time()
    with _map_lock:
        fresh = _map["gb2reg"] and (now - _map["ts"]) < MAP_TTL
        if fresh and not force:
            return _map["gb2reg"], _map["reg2gb"], _map["regs"]
    gb2reg, reg2gb, regs = _build_map()
    with _map_lock:
        _map.update(ts=now, gb2reg=gb2reg, reg2gb=reg2gb, regs=regs)
    return gb2reg, reg2gb, regs


# ==================== VISOKA RAZINA: lookup za /gdje ====================

def _disp_reg(reg_norm, regs):
    """Vrati originalni (citljivi) oblik registracije za normalizirani kljuc."""
    for r in regs:
        if norm_reg(r) == reg_norm:
            return r
    return reg_norm


def _suggest(query, pool, n=3):
    """2-3 najslicnije registracije (po slicnosti normaliziranih oblika)."""
    import difflib
    qn = norm_reg(query)
    uniq, seen = [], set()
    for r in pool:
        rn = norm_reg(r)
        if rn and rn not in seen:
            seen.add(rn)
            uniq.append(r)
    uniq.sort(key=lambda r: difflib.SequenceMatcher(None, qn, norm_reg(r)).ratio(),
              reverse=True)
    return uniq[:n]


def lookup(query):
    """Glavna funkcija za /gdje. Vrati dict sa 'status':
      ok         -> reg, gb, pos
      no_device  -> reg, gb (vozilo u Excelu, ali nema GPS uredaj)
      no_position-> reg, gb (uredaj postoji, ali nema trenutne pozicije)
      not_found  -> query, suggestions
      empty / error -> (poruka)
    Ne baca iznimke — greske vraca kao status='error'."""
    q = _txt(query)
    if not q:
        return {"status": "empty"}
    if not is_configured():
        return {"status": "error",
                "message": "Mobilisis nije konfiguriran (MOBILISIS_USER/PASS)."}

    try:
        gb2reg, reg2gb, regs = _get_map()
    except Exception as e:
        _log(f"mapiranje GB/REG nije uspjelo: {e}")
        monitoring.error("Mobilisis: mapiranje GB/REG nije uspjelo",
                         source="mobilisis", exc=e)
        return {"status": "error",
                "message": "Ne mogu učitati GARAŽNI BROJEVI.xlsx (GB↔REG)."}

    # Upit = samo znamenke -> GB; inace registracija
    if re.fullmatch(r"\d+", q):
        gb = norm_gb(q)
        reg_norm = gb2reg.get(gb)
        if not reg_norm:
            return {"status": "not_found", "query": q,
                    "suggestions": _suggest(q, regs)}
    else:
        reg_norm = norm_reg(q)
        gb = reg2gb.get(reg_norm)

    try:
        devices = get_devices()
    except Exception as e:
        _log(f"get_devices greška: {e}")
        monitoring.error("Mobilisis: get_devices greška", source="mobilisis", exc=e)
        return {"status": "error", "message": "Mobilisis API ne odgovara (uređaji)."}

    by_norm = {}
    for d in devices:
        by_norm.setdefault(norm_reg(d["Name"]), d)
    dev = by_norm.get(reg_norm) if reg_norm else None

    if dev is None:
        # Vozilo poznato iz Excela (reg ili gb), ali nema GPS uredaja?
        if reg_norm and (reg_norm in reg2gb or gb):
            return {"status": "no_device",
                    "reg": _disp_reg(reg_norm, regs), "gb": gb}
        return {"status": "not_found", "query": q,
                "suggestions": _suggest(q, [d["Name"] for d in devices] + regs)}

    reg_display = dev["Name"] or _disp_reg(reg_norm, regs)
    try:
        positions = get_positions([dev["Id"]])
    except Exception as e:
        _log(f"get_positions greška: {e}")
        monitoring.error("Mobilisis: get_positions greška", source="mobilisis", exc=e)
        return {"status": "error", "message": "Mobilisis API ne odgovara (pozicije)."}

    pos = positions[0] if positions else None
    if not pos or pos.get("lat") is None or pos.get("lon") is None:
        return {"status": "no_position", "reg": reg_display, "gb": gb}
    return {"status": "ok", "reg": reg_display, "gb": gb, "pos": pos}


# ==================== SVE POZICIJE (za HTTP endpoint) ====================

def all_positions():
    """Sve trenutne pozicije svih vozila, obogacene s registracijom i GB.

    Vraca listu dictova (JSON-spremnih):
      {gb, registracija, lat, lon, brzina, smjer, motor(bool),
       vrijeme(ISO8601 UTC ili None), odometar}
    lat/lon/brzina/smjer/odometar su brojevi (ili None). Vozila bez GB
    mapiranja dobiju gb=None. Baca MobilisisError na gresku (login/API/mapiranje)."""
    gb2reg, reg2gb, regs = _get_map()
    devices = get_devices()
    id2name = {str(d["Id"]): d.get("Name", "") for d in devices}
    positions = get_positions()

    out = []
    for p in positions:
        did = str(p.get("deviceId"))
        reg_display = _txt(id2name.get(did, ""))
        reg_norm = norm_reg(reg_display)
        gb = reg2gb.get(reg_norm) if reg_norm else None
        dt = parse_utc(p.get("dateTime"))
        out.append({
            "gb": gb or None,
            "registracija": reg_display or None,
            "lat": p.get("lat"),
            "lon": p.get("lon"),
            "brzina": p.get("speed"),
            "smjer": p.get("heading"),
            "motor": ignition_on(p.get("ignition")),
            "vrijeme": dt.astimezone(timezone.utc).isoformat() if dt else None,
            "odometar": p.get("odometer"),
        })
    return out
