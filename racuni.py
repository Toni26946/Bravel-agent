# ============================================================
#  DOKUMENTI (racuni + primke) - obrada fotografija s Telegrama
#  i upis u Excel na SharePointu (preko graph_client / MS Graph).
#
#  Tok:
#   1. vozac posalje fotku (photo ili slika-kao-dokument); za
#      visestranicne primke posalje album ili vise fotki + /gotovo
#   2. Claude vision (claude-sonnet-4-6) KLASIFICIRA vrstu
#      ("racun" -> Racuni_terena.xlsx / "primka" -> Primke_terena.xlsx)
#      i izvuce podatke -> JSON (svi listovi u JEDNOM pozivu)
#   3. mapiramo vozaca/zaprimioca (SQLite); za primku pitamo GB
#   4. sazetak s vrstom + gumbi ✅ Upiši / ✏️ Ispravi / ❌ Odbaci /
#      🔄 Promijeni vrstu
#   5. tek nakon ✅ upisujemo retke (redak = jedna stavka)
#
#  Integracija: main.py poziva setup(), init_db(), register(),
#  a na vrhu svog catch-all text handlera poziva handle_text(message).
# ============================================================

import base64
import io
import json
import re
import threading
import time
from datetime import datetime

import telebot

import graph_client
import monitoring

# ---- Postavke ----
VISION_MODEL = "claude-sonnet-4-6"  # citanje/klasifikacija; razgovor ostaje Haiku


class DocSpec:
    """Opis vrste dokumenta: gdje se upisuje i po kojim kolonama."""

    def __init__(self, vrsta, emoji, naziv, sheet_label, excel_file, table,
                 columns, new_markers, key_data, key_cols):
        self.vrsta = vrsta            # 'racun' | 'primka'
        self.emoji = emoji
        self.naziv = naziv            # 'Račun' / 'Primka'
        self.sheet_label = sheet_label  # 'Racuni_terena' / 'Primke_terena'
        self.excel_file = excel_file
        self.table = table
        self.columns = columns
        self.new_markers = new_markers  # kolone koje oznacavaju NOVU strukturu
        self.key_data = key_data        # kljucevi u data dictu (oib, broj)
        self.key_cols = key_cols        # kolone u tablici (OIB, Broj) za dedupe


RACUN = DocSpec(
    vrsta="racun", emoji="🧾", naziv="Račun", sheet_label="Racuni_terena",
    excel_file="Racuni_terena.xlsx", table="Racuni",
    columns=[
        "Datum", "Vrijeme", "Izdavatelj", "OIB", "BrojRacuna",
        "Stavka", "Kolicina", "JedinicnaCijena", "IznosStavke",
        "PDV", "UkupnoEUR", "Lokacija", "Vozac", "GB",
        "VrijemeUnosa", "UnioTelegramID",
    ],
    new_markers=("Stavka", "IznosStavke", "Lokacija"),
    key_data=("oib", "broj_racuna"),
    key_cols=("OIB", "BrojRacuna"),
)

PRIMKA = DocSpec(
    vrsta="primka", emoji="📦", naziv="Primka", sheet_label="Primke_terena",
    excel_file="Primke_terena.xlsx", table="Primke",
    columns=[
        "DatumDokumenta", "Firma", "OIB", "BrRacuna", "KataloskiBroj",
        "Stavka", "Kolicina", "JM", "JedinicnaCijena", "RabatPosto",
        "IznosStavke", "UkupnoEUR", "Dospijece", "GB", "Zaprimio",
        "VrijemeUnosa", "UnioTelegramID",
    ],
    new_markers=("KataloskiBroj", "Dospijece", "Zaprimio"),
    key_data=("oib", "br_racuna"),
    key_cols=("OIB", "BrRacuna"),
)


def _spec_for(vrsta):
    return PRIMKA if vrsta == "primka" else RACUN


# ---- Ovisnosti koje ubrizgava main.py preko setup() ----
_bot = None
_client = None            # anthropic klijent
_db = None                # tvornica konekcija (main.db)
_allowed_users = ()
_tz = None
_log_note = None          # main.log_note (opcionalno biljezenje u dnevni log)

# ---- Stanje razgovora po chatu (jedan aktivan dokument po korisniku) ----
_sessions = {}
_sessions_lock = threading.Lock()
_token_counter = [0]      # rastuci token za callback_data (kratak, <64 B)

# ---- Skupljanje visestranicnih dokumenata (album / vise fotki) ----
_collectors = {}
_collectors_lock = threading.Lock()


def setup(bot, client, db, allowed_users, tz, log_note=None):
    global _bot, _client, _db, _allowed_users, _tz, _log_note
    _bot = bot
    _client = client
    _db = db
    _allowed_users = tuple(allowed_users)
    _tz = tz
    _log_note = log_note


def _now():
    return datetime.now(_tz)


def _log(msg):
    print(f"[racuni] {msg}", flush=True)


# ==================== BAZA: mapiranje vozaca ====================

def init_db():
    """Kreira tablicu vozaca u glavnoj bazi (bot.db)."""
    with _db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS vozaci (
                telegram_user_id INTEGER PRIMARY KEY,
                ime_vozaca       TEXT NOT NULL,
                gb_vozila        TEXT
            )
        """)
    print("Tablica vozaca spremna.")


def get_driver(user_id):
    """Vraca (ime, gb) ili None ako korisnik nije mapiran."""
    with _db() as conn:
        row = conn.execute(
            "SELECT ime_vozaca, gb_vozila FROM vozaci WHERE telegram_user_id = ?",
            (user_id,)
        ).fetchone()
    if row:
        return row["ime_vozaca"], row["gb_vozila"]
    return None


def _safe_get_driver(user_id):
    try:
        return get_driver(user_id)
    except Exception as e:
        monitoring.warning(f"Racuni: dohvat vozaca nije uspio: {e}", source="racuni")
        return None


def upsert_driver(user_id, ime, gb):
    with _db() as conn:
        conn.execute(
            "INSERT INTO vozaci (telegram_user_id, ime_vozaca, gb_vozila) "
            "VALUES (?, ?, ?) "
            "ON CONFLICT(telegram_user_id) DO UPDATE SET "
            "ime_vozaca = excluded.ime_vozaca, gb_vozila = excluded.gb_vozila",
            (user_id, ime, gb)
        )


def list_drivers():
    with _db() as conn:
        return conn.execute(
            "SELECT telegram_user_id, ime_vozaca, gb_vozila FROM vozaci "
            "ORDER BY ime_vozaca"
        ).fetchall()


# ==================== POMOCNE ====================

def _allowed(message_or_user_id):
    uid = (message_or_user_id.from_user.id
           if hasattr(message_or_user_id, "from_user") else message_or_user_id)
    return uid in _allowed_users


def _parse_num(val):
    """Pretvori '1.234,56' / '1234.56' / broj u float, ili None ako ne ide."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s:
        return None
    # Ako ima i tocku i zarez -> zarez je decimalni (hr format): makni tocke
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_num(x):
    """Broj za prikaz (2 decimale) ili '⚠️' ako fali."""
    if x is None:
        return "⚠️"
    return f"{x:.2f}"


def _txt(v):
    return str(v).strip() if v not in (None, "") else ""


# ==================== CLAUDE VISION (klasifikacija + citanje) ====================

_VISION_SYSTEM = (
    "Ti si precizan sustav za citanje hrvatskih poslovnih dokumenata sa slike "
    "(fiskalni racuni i veleprodajne primke / racun-otpremnice). Odgovaraj "
    "ISKLJUCIVO validnim JSON-om, bez ikakvog teksta prije ili poslije."
)


def _vision_prompt(force_vrsta=None):
    klas = (
        "Prvo KLASIFICIRAJ dokument u polje \"vrsta\":\n"
        "  - \"racun\": maloprodajni FISKALNI racun (ima JIR i/ili ZKI, kratke "
        "stavke bez sifri artikala).\n"
        "  - \"primka\": veleprodajni racun-otpremnica / primka (stavke sa "
        "SIFROM artikla, rabat, dospijece, kupac naveden na dokumentu).\n"
    )
    if force_vrsta == "racun":
        klas = "Tretiraj dokument kao \"racun\" (fiskalni racun).\n"
    elif force_vrsta == "primka":
        klas = "Tretiraj dokument kao \"primka\" (veleprodajna primka/otpremnica).\n"

    return (
        "Slike su STRANICE JEDNOG dokumenta (moze ih biti vise). " + klas +
        "Ako na dokumentu pise \"Stranica X/Y\", u \"broj_stranica\" vrati Y "
        "(ukupan broj stranica), inace 1.\n\n"
        "Ako je vrsta \"racun\", vrati STROGO JSON:\n"
        '{\n'
        '  "vrsta": "racun",\n'
        '  "broj_stranica": broj,\n'
        '  "datum": "DD.MM.YYYY ili null",\n'
        '  "vrijeme": "HH:MM ili null",\n'
        '  "izdavatelj": "naziv tvrtke ili null",\n'
        '  "oib": "OIB izdavatelja (11 znamenki) ili null",\n'
        '  "broj_racuna": "broj racuna ili null",\n'
        '  "lokacija": "adresa/mjesto poslovnice iz zaglavlja ili null",\n'
        '  "stavke": [ {"naziv": "...", "kolicina": broj_ili_null, '
        '"cijena": broj_ili_null, "iznos": broj_ili_null} ],\n'
        '  "ukupno_eur": broj_ili_null,\n'
        '  "pdv_iznos": broj_ili_null\n'
        '}\n\n'
        "Ako je vrsta \"primka\", vrati STROGO JSON:\n"
        '{\n'
        '  "vrsta": "primka",\n'
        '  "broj_stranica": broj,\n'
        '  "datum_dokumenta": "DD.MM.YYYY ili null",\n'
        '  "firma": "naziv dobavljaca/izdavatelja ili null",\n'
        '  "oib": "OIB izdavatelja (11 znamenki) ili null",\n'
        '  "br_racuna": "PUNA oznaka dokumenta ili null",\n'
        '  "dospijece": "DD.MM.YYYY ili null",\n'
        '  "stavke": [ {"kataloski_broj": "...", "naziv": "...", '
        '"kolicina": broj_ili_null, "jm": "kom/kg/l...", '
        '"jedinicna_cijena": broj_ili_null, "rabat_posto": broj_ili_null, '
        '"iznos": broj_ili_null} ],\n'
        '  "ukupno_eur": broj_ili_null\n'
        '}\n\n'
        "PRAVILA za primku:\n"
        "- \"br_racuna\": PUNA oznaka dokumenta iz naslova, npr. iz "
        "\"Racun - Otpremnica 5795-03-261\" vrati \"5795-03-261\" "
        "(NE skraceni \"Broj dokumenta\").\n"
        "- \"kataloski_broj\": sifra artikla iz kolone \"Šifra i naziv "
        "artikla\" (npr. 038103385A) — NE barkod/EAN kolona.\n"
        "- skupi stavke sa SVIH stranica.\n\n"
        "Sve novcane iznose vrati kao brojeve u eurima (npr. 12.50). Ako polje "
        "ne postoji ili je necitljivo, stavi null (nemoj izmisljati). "
        "Vrati SAMO JSON."
    )


def _extract_json(text):
    """Izvuci JSON objekt iz Claudeova odgovora (skini ograde, nadji {...})."""
    t = text.strip()
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?\s*", "", t)
        t = re.sub(r"\s*```$", "", t)
    start = t.find("{")
    end = t.rfind("}")
    if start != -1 and end != -1 and end > start:
        t = t[start:end + 1]
    return json.loads(t)


def _read_document(images, force_vrsta=None):
    """images: lista (bytes, media_type). Posalji SVE stranice u jednom pozivu
    Claude visionu i vrati parsirani dict (ukljucivo 'vrsta')."""
    content = []
    for (b, mt) in images:
        b64 = base64.standard_b64encode(b).decode("utf-8")
        content.append({"type": "image", "source": {
            "type": "base64", "media_type": mt, "data": b64}})
    content.append({"type": "text", "text": _vision_prompt(force_vrsta)})

    resp = _client.messages.create(
        model=VISION_MODEL,
        max_tokens=2500,
        system=_VISION_SYSTEM,
        messages=[{"role": "user", "content": content}],
        temperature=0,
    )
    text = next((b.text for b in resp.content if b.type == "text"), "")
    data = _extract_json(text)
    if force_vrsta:
        data["vrsta"] = force_vrsta
    if data.get("vrsta") not in ("racun", "primka"):
        data["vrsta"] = "racun"
    return data


# ==================== STAVKE ====================

def _stavka_usable(s):
    """True ako stavka ima barem naziv ili citljiv iznos."""
    if not isinstance(s, dict):
        return False
    naziv = (s.get("naziv") or "").strip()
    return bool(naziv) or _parse_num(s.get("iznos")) is not None


def _usable_stavke(data):
    return [s for s in (data.get("stavke") or []) if _stavka_usable(s)]


def _racun_stavka_fields(s):
    """(naziv, kolicina, jed_cijena, iznos) — brojevi ili None gdje fali."""
    if not isinstance(s, dict):
        return ("(nespecificirano)", None, None, None)
    naziv = (s.get("naziv") or "").strip() or "(nespecificirano)"
    return (naziv, _parse_num(s.get("kolicina")),
            _parse_num(s.get("cijena")), _parse_num(s.get("iznos")))


def _primka_stavka_fields(s):
    """(kataloski, naziv, kolicina, jm, jed_cijena, rabat, iznos)."""
    if not isinstance(s, dict):
        return ("", "(nespecificirano)", None, "", None, None, None)
    naziv = (s.get("naziv") or "").strip() or "(nespecificirano)"
    return (_txt(s.get("kataloski_broj")), naziv,
            _parse_num(s.get("kolicina")), _txt(s.get("jm")),
            _parse_num(s.get("jedinicna_cijena")),
            _parse_num(s.get("rabat_posto")), _parse_num(s.get("iznos")))


def _racun_stavka_line(s):
    naziv, kol, cij, izn = _racun_stavka_fields(s)
    kol_s = f"{kol:g}" if kol is not None else "⚠️"
    cij_s = f"{cij:.2f}" if cij is not None else "⚠️"
    izn_s = f"{izn:.2f}" if izn is not None else "⚠️"
    return f"  • {naziv} — {kol_s} × {cij_s} = {izn_s}"


def _primka_stavka_line(s):
    kat, naziv, kol, jm, cij, rab, izn = _primka_stavka_fields(s)
    pref = f"[{kat}] " if kat else ""
    kol_s = f"{kol:g}" if kol is not None else "⚠️"
    jm_s = f" {jm}" if jm else ""
    cij_s = f"{cij:.2f}" if cij is not None else "⚠️"
    rab_s = f" (-{rab:g}%)" if rab else ""
    izn_s = f"{izn:.2f}" if izn is not None else "⚠️"
    return f"  • {pref}{naziv} — {kol_s}{jm_s} × {cij_s}{rab_s} = {izn_s}"


# ==================== SAZETAK I GUMBI ====================

def _summary_text(sess):
    data = sess["data"]
    spec = sess["spec"]

    def show(k):
        v = data.get(k)
        return str(v).strip() if v not in (None, "") else "⚠️"

    lines = [f"{spec.emoji} {spec.naziv} → {spec.sheet_label}", ""]
    usable = _usable_stavke(data)

    if sess["vrsta"] == "primka":
        lines.append(f"• Datum dokumenta: {show('datum_dokumenta')}")
        lines.append(f"• Firma: {show('firma')}")
        lines.append(f"• OIB: {show('oib')}")
        lines.append(f"• Br. računa: {show('br_racuna')}")
        lines.append(f"• Dospijeće: {show('dospijece')}")
        lines.append("")
        lines.append("Stavke:")
        if usable:
            for s in usable:
                lines.append(_primka_stavka_line(s))
        else:
            lines.append("  ⚠️ nema čitljivih stavki — upisat će se 1 redak "
                         "„(nespecificirano)” s ukupnim iznosom")
        lines.append("")
        lines.append(f"• Ukupno (EUR): {_fmt_num(_parse_num(data.get('ukupno_eur')))}")
        lines.append(f"• GB (vozilo): {sess.get('gb') or '⚠️'}")
        lines.append(f"• Zaprimio: {sess.get('zaprimio') or '⚠️'}")
    else:
        lines.append(f"• Datum: {show('datum')}   Vrijeme: {show('vrijeme')}")
        lines.append(f"• Izdavatelj: {show('izdavatelj')}")
        lines.append(f"• OIB: {show('oib')}")
        lines.append(f"• Broj računa: {show('broj_racuna')}")
        lines.append(f"• Lokacija: {show('lokacija')}")
        lines.append("")
        lines.append("Stavke:")
        if usable:
            for s in usable:
                lines.append(_racun_stavka_line(s))
        else:
            lines.append("  ⚠️ nema čitljivih stavki — upisat će se 1 redak "
                         "„(nespecificirano)” s ukupnim iznosom")
        lines.append("")
        lines.append(f"• PDV: {_fmt_num(_parse_num(data.get('pdv_iznos')))}   "
                     f"Ukupno (EUR): {_fmt_num(_parse_num(data.get('ukupno_eur')))}")
        lines.append(f"• Vozač: {sess.get('vozac') or '⚠️'}")
        lines.append(f"• GB (vozilo): {sess.get('gb') or '⚠️'}")

    lines.append("")
    n = len(usable) if usable else 1
    lines.append(f"➡️ Upisat će se {n} redak(a) — jedan po stavci.")
    lines.append("⚠️ = nečitljivo / prazno. Ispravi ili promijeni vrstu po potrebi.")
    return "\n".join(lines)


def _confirm_markup(token):
    mk = telebot.types.InlineKeyboardMarkup()
    mk.row(
        telebot.types.InlineKeyboardButton("✅ Upiši", callback_data=f"rc_ok_{token}"),
        telebot.types.InlineKeyboardButton("✏️ Ispravi", callback_data=f"rc_ed_{token}"),
        telebot.types.InlineKeyboardButton("❌ Odbaci", callback_data=f"rc_no_{token}"),
    )
    mk.row(
        telebot.types.InlineKeyboardButton("🔄 Promijeni vrstu",
                                           callback_data=f"rc_ty_{token}"),
    )
    return mk


def _send_confirm(sess):
    _bot.send_message(sess["chat_id"], _summary_text(sess),
                      reply_markup=_confirm_markup(sess["token"]))


# ==================== EXCEL: gradnja redaka ====================

def _num(x):
    return x if x is not None else ""  # broj ili prazna celija


def _build_racun_rows(sess):
    data = sess["data"]
    ukupno = _parse_num(data.get("ukupno_eur"))
    pdv = _parse_num(data.get("pdv_iznos"))
    head = [_txt(data.get("datum")), _txt(data.get("vrijeme")),
            _txt(data.get("izdavatelj")), _txt(data.get("oib")),
            _txt(data.get("broj_racuna"))]                       # 5 kolona
    tail = [_num(pdv), _num(ukupno), _txt(data.get("lokacija")),
            sess.get("vozac") or "", sess.get("gb") or "",
            _now().strftime("%Y-%m-%d %H:%M:%S"),
            str(sess["user_id"])]                                # 7 kolona
    usable = _usable_stavke(data)
    rows = []
    if usable:
        for s in usable:
            naziv, kol, cij, izn = _racun_stavka_fields(s)
            rows.append(head + [naziv, _num(kol), _num(cij), _num(izn)] + tail)
    else:
        rows.append(head + ["(nespecificirano)", 1, _num(ukupno), _num(ukupno)]
                    + tail)
    return rows


def _build_primka_rows(sess):
    data = sess["data"]
    ukupno = _parse_num(data.get("ukupno_eur"))
    head = [_txt(data.get("datum_dokumenta")), _txt(data.get("firma")),
            _txt(data.get("oib")), _txt(data.get("br_racuna"))]  # 4 kolone
    tail = [_num(ukupno), _txt(data.get("dospijece")),
            sess.get("gb") or "", sess.get("zaprimio") or "",
            _now().strftime("%Y-%m-%d %H:%M:%S"),
            str(sess["user_id"])]                                # 6 kolona
    usable = _usable_stavke(data)
    rows = []
    if usable:
        for s in usable:
            kat, naziv, kol, jm, cij, rab, izn = _primka_stavka_fields(s)
            rows.append(head + [kat, naziv, _num(kol), jm, _num(cij),
                                _num(rab), _num(izn)] + tail)     # 7 kolona
    else:
        rows.append(head + ["", "(nespecificirano)", 1, "", _num(ukupno),
                            "", _num(ukupno)] + tail)
    return rows


def _build_rows(sess):
    if sess["vrsta"] == "primka":
        return _build_primka_rows(sess)
    return _build_racun_rows(sess)


# ==================== EXCEL: kreiranje / append (generic) ====================

def _col_letter(n):
    """1 -> A, 16 -> P."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _create_workbook_bytes(spec, initial_rows=None):
    """Kreiraj novi xlsx s Excel TABLICOM (ListObject). Prvi upis NE ovisi o
    workbook API-ju na tek nastalom fajlu."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = spec.table
    ws.append(spec.columns)
    for r in (initial_rows or []):
        ws.append(r)

    last_col = _col_letter(len(spec.columns))
    tab = Table(displayName=spec.table, ref=f"A1:{last_col}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_is_empty(values):
    """True ako su SVE vrijednosti retka None ili prazan string."""
    return all(v is None or str(v).strip() == "" for v in values)


def _fallback_append(spec, rows):
    """Rezerva ako workbook API zapne: download -> KOMPAKTIRAJ+dopisi -> upload.

    Ne oslanjamo se na ws.append()/ws.max_row (broje i prazne formatirane
    celije). Procitamo sve NEPRAZNE retke, dopisemo nove i upisemo kompaktno
    od retka 2. Novi redak zavrsi neposredno iza zadnjeg nepraznog; prazni
    'duh' redci (trailing/unutarnji) se ciste; raspon tablice se stisne."""
    from openpyxl import load_workbook

    content = graph_client.download_file(spec.excel_file)
    wb = load_workbook(io.BytesIO(content))
    ws = wb[spec.table] if spec.table in wb.sheetnames else wb.active
    n_cols = len(spec.columns)
    old_max = ws.max_row

    data = []
    for r in range(2, old_max + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, n_cols + 1)]
        if not _row_is_empty(vals):
            data.append(vals)
    for r in rows:
        data.append([(r[j] if j < len(r) else None) for j in range(n_cols)])

    for i, vals in enumerate(data):
        for c in range(n_cols):
            ws.cell(row=2 + i, column=c + 1, value=vals[c])
    last_row = 1 + len(data)
    for r in range(last_row + 1, old_max + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).value = None

    tab = ws.tables.get(spec.table) if hasattr(ws, "tables") else None
    if tab is not None:
        tab.ref = f"A1:{_col_letter(n_cols)}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    graph_client.upload_file(spec.excel_file, buf.getvalue())


# HTTP kodovi za KRATKI, odmah ponovljeni pokusaj.
_RETRYABLE = {404, 500, 502, 503, 504}
# Kodovi "fajl otvoren/zakljucan / konflikt / rate-limit" -> DUGI backoff.
_LOCK_CODES = {423, 409, 429}
_LOCK_BACKOFF = [15, 30, 60, 120, 240]


class _Locked(Exception):
    """Fajl je otvoren/zakljucan (423) ili konflikt/rate-limit (409/429)."""

    def __init__(self, status):
        super().__init__(f"locked (HTTP {status})")
        self.status = status


def _append_via_workbook(spec, rows):
    """Dodaj retke preko workbook API-ja s kratkim backoff retryjem.
    Koristi append_or_fill: popuni prazne 'duh' retke (PATCH), ostatak rows/add
    — imuno na prazne retke od rucnog brisanja. Dize GraphError ako ne uspije."""
    delays = [2, 4]
    for attempt in range(3):
        try:
            graph_client.append_or_fill_table_rows(spec.excel_file, spec.table, rows)
            return
        except graph_client.GraphError as e:
            if e.status_code in _RETRYABLE and attempt < len(delays):
                time.sleep(delays[attempt])
                continue
            raise


# ---- Migracija starog fajla (stara struktura -> _STARO) ----

def _read_header(spec):
    from openpyxl import load_workbook

    content = graph_client.download_file(spec.excel_file)
    wb = load_workbook(io.BytesIO(content), read_only=True)
    try:
        ws = wb[spec.table] if spec.table in wb.sheetnames else wb.active
        first = next(ws.iter_rows(min_row=1, max_row=1), None)
        return [c.value for c in first] if first else []
    finally:
        wb.close()


def _is_old_structure(spec):
    """True ako postojeci fajl NEMA nove kolone (stara struktura)."""
    try:
        header = _read_header(spec)
    except Exception as e:
        _log(f"provjera strukture nije uspjela ({e}); pretpostavljam novu")
        return False
    if not header:
        return False
    return not all(marker in header for marker in spec.new_markers)


def _rename_old_aside(spec):
    """Preimenuj stari fajl u <ime>_STARO.xlsx (uz sufiks ako je ime zauzeto)."""
    base = spec.excel_file[:-5] if spec.excel_file.endswith(".xlsx") else spec.excel_file
    base = base + "_STARO"
    ext = ".xlsx"
    candidates = [f"{base}{ext}"] + [f"{base}_{i}{ext}" for i in range(2, 21)]
    for name in candidates:
        try:
            graph_client.rename_file(spec.excel_file, name)
            return name
        except graph_client.GraphError as e:
            if e.status_code == 409:
                continue
            raise
    name = f"{base}_{_now().strftime('%Y%m%d_%H%M%S')}{ext}"
    graph_client.rename_file(spec.excel_file, name)
    return name


# ==================== DEDUPE (OIB + broj dokumenta) ====================

def _norm_key(v):
    """Normaliziraj vrijednost kljuca u string. Excel zna procitati broj kao
    float (2700809.0) — skidamo '.0'."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        core = s[:-2]
        if core.lstrip("-").isdigit():
            return core
    return s


def _receipt_key(spec, data):
    return tuple(_norm_key(data.get(k)) for k in spec.key_data)


def _existing_keys(spec):
    """Svjez download pravog fajla; vrati set (oib, broj) postojecih redaka —
    normalizirano, ignorirajuci prazne retke. Prazan set ako fajl/kolone fale."""
    if not graph_client.file_exists(spec.excel_file):
        return set()
    from openpyxl import load_workbook

    content = graph_client.download_file(spec.excel_file)  # svjez, ne kesira se
    wb = load_workbook(io.BytesIO(content), read_only=True)
    try:
        ws = wb[spec.table] if spec.table in wb.sheetnames else wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return set()
        header = list(header)
        oib_col, num_col = spec.key_cols
        if oib_col not in header or num_col not in header:
            return set()
        i_oib = header.index(oib_col)
        i_num = header.index(num_col)
        keys = set()
        for r in it:
            if _row_is_empty(r):
                continue
            oib = _norm_key(r[i_oib]) if i_oib < len(r) else ""
            num = _norm_key(r[i_num]) if i_num < len(r) else ""
            if num:  # kljuc smislen samo ako ima broj dokumenta
                keys.add((oib, num))
        return keys
    finally:
        wb.close()


def _is_duplicate(spec, data):
    """True ako dokument s istim (OIB, broj) vec postoji. Bez broja ne dedupe."""
    key = _receipt_key(spec, data)
    if not key[-1]:
        return False
    return key in _existing_keys(spec)


def _dup_warning_text(sess):
    data = sess["data"]
    spec = sess["spec"]
    num = data.get(spec.key_data[1])
    return (f"⚠️ Ova {spec.naziv.lower()} je već upisana!\n"
            f"OIB: {data.get('oib') or '?'}   Broj: {num or '?'}\n\n"
            "Želiš li je svejedno upisati?")


def _dup_markup(token):
    mk = telebot.types.InlineKeyboardMarkup()
    mk.row(
        telebot.types.InlineKeyboardButton("✅ Ipak upiši", callback_data=f"rc_fo_{token}"),
        telebot.types.InlineKeyboardButton("❌ Odustani", callback_data=f"rc_no_{token}"),
    )
    return mk


# ==================== UPIS ====================

def _write_once(sess):
    """JEDAN pokusaj upisa. Vraca (ok, poruka). Dize _Locked ako je fajl
    zakljucan/konflikt (423/409/429)."""
    spec = sess["spec"]
    if not graph_client.is_configured():
        _log("SharePoint nije konfiguriran (nedostaju Graph kredencijali).")
        return False, ("SharePoint nije konfiguriran (nedostaju Graph "
                       "kredencijali). Redak nije upisan.")

    step = "start"
    try:
        step = "1/6 auth"
        graph_client.ensure_token()

        rows = _build_rows(sess)
        _log(f"[{spec.vrsta}] pripremljeno redaka za upis: {len(rows)}")

        step = "2/6 provjera fajla"
        exists = graph_client.file_exists(spec.excel_file)
        _log(f"2/6 {spec.excel_file} postoji = {exists}")

        if exists:
            step = "2b/6 provjera strukture"
            if _is_old_structure(spec):
                old = _rename_old_aside(spec)
                _log(f"2b/6 STARA struktura -> preimenovano u {old}; kreiram novi")
                monitoring.info(
                    f"{spec.naziv}: stari fajl preimenovan u {old} (migracija).",
                    source="racuni")
                exists = False

        if not exists:
            step = "3/6 kreiranje xlsx"
            content = _create_workbook_bytes(spec, initial_rows=rows)
            step = "4/6 upload"
            graph_client.upload_file(spec.excel_file, content)
            _log("4/6 upload OK")
            return True, (f"✅ {spec.naziv} upisana — {len(rows)} stavka(i) "
                          f"(kreiran novi {spec.excel_file}).")

        step = "5/6 workbook append"
        _log(f"5/6 workbook append {len(rows)} redaka (append_or_fill)")
        try:
            _append_via_workbook(spec, rows)
            _log("5/6 workbook append OK")
            return True, (f"✅ {spec.naziv} upisana — {len(rows)} stavka(i) u "
                          f"{spec.excel_file}.")
        except graph_client.GraphError as e:
            if e.status_code in _LOCK_CODES:
                raise
            step = "5/6 fallback download/upload"
            _log(f"5/6 workbook append pao ({e}); fallback download->upload")
            monitoring.warning(
                f"Workbook API zapeo ({e}); prelazim na fallback.",
                source="racuni")
            _fallback_append(spec, rows)
            _log("5/6 fallback OK")
            return True, (f"✅ {spec.naziv} upisana — {len(rows)} stavka(i) "
                          "(rezervna metoda download→upload).")

    except graph_client.GraphError as e:
        if e.status_code in _LOCK_CODES:
            _log(f"{step} zakljucano/konflikt (HTTP {e.status_code})")
            raise _Locked(e.status_code)
        _log(f"{step} GREŠKA: GraphError kod={e.status_code} {e}")
        if e.status_code == 403:
            monitoring.error("Graph 403 pri upisu", source="racuni", exc=e)
            return False, ("❌ Nedostaje ovlast — provjeri admin consent za "
                           "Sites.ReadWrite.All.")
        monitoring.error("Graph greska pri upisu", source="racuni", exc=e)
        return False, f"❌ Greška pri upisu na SharePoint (kod {e.status_code})."
    except Exception as e:
        _log(f"{step} GREŠKA: {type(e).__name__}: {e}")
        monitoring.error("Neocekivana greska pri upisu", source="racuni", exc=e)
        return False, "❌ Neočekivana greška pri upisu dokumenta."


# ==================== SKUPLJANJE VISESTRANICNIH DOKUMENATA ====================

def _download_telegram_file(file_id):
    info = _bot.get_file(file_id)
    return _bot.download_file(info.file_path)


def _obradi_markup():
    mk = telebot.types.InlineKeyboardMarkup()
    mk.row(telebot.types.InlineKeyboardButton(
        "📄 Obradi dokument", callback_data="doc_go"))
    return mk


def _collector_text(n, total=None, first=False):
    if total and total > 1 and first:
        head = f"📄 Stranica 1/{total} — pošalji ostale stranice"
    else:
        head = f"📄 Primljeno {n} stranica — pošalji ostale"
    return (f"{head} pa klikni „Obradi dokument”.\n"
            "(ili napiši /gotovo)")


def _ingest_image(message, img, mtype):
    """Photo/dokument -> ili u aktivni kolektor, ili novi kolektor (album),
    ili obrada odmah (standalone)."""
    chat_id = message.chat.id
    with _collectors_lock:
        col = _collectors.get(chat_id)
        if col is not None:
            col["images"].append((img, mtype))
            created = False
            msg_id = col.get("msg_id")
            n = len(col["images"])
        elif message.media_group_id:
            col = {"images": [(img, mtype)], "msg_id": None,
                   "who": message.from_user.full_name,
                   "media_group_id": message.media_group_id}
            _collectors[chat_id] = col
            created = True
            msg_id = None
            n = 1
        else:
            col = None

    if col is None:
        # standalone: obradi odmah (u threadu)
        threading.Thread(
            target=_process_standalone,
            args=(chat_id, message.from_user.id, message.from_user.full_name,
                  img, mtype),
            daemon=True).start()
        return

    if created:
        # tvorac kolektora salje prompt s gumbom (samo jednom)
        try:
            m = _bot.send_message(chat_id, _collector_text(n),
                                  reply_markup=_obradi_markup())
            with _collectors_lock:
                c2 = _collectors.get(chat_id)
                if c2 is not None and c2.get("msg_id") is None:
                    c2["msg_id"] = m.message_id
        except Exception:
            pass
    elif msg_id is not None:
        # ostale stranice: best-effort osvjezi brojac (bez dupliciranja poruke)
        try:
            _bot.edit_message_text(_collector_text(n), chat_id, msg_id,
                                   reply_markup=_obradi_markup())
        except Exception:
            pass


def _start_collector_for_multipage(chat_id, who, img, mtype, total):
    """Standalone fotka za koju je vision javio Y>1 -> zapocni kolektor."""
    with _collectors_lock:
        if chat_id not in _collectors:
            _collectors[chat_id] = {"images": [(img, mtype)], "msg_id": None,
                                    "who": who, "media_group_id": None}
    try:
        m = _bot.send_message(chat_id, _collector_text(1, total=total, first=True),
                              reply_markup=_obradi_markup())
        with _collectors_lock:
            c2 = _collectors.get(chat_id)
            if c2 is not None and c2.get("msg_id") is None:
                c2["msg_id"] = m.message_id
    except Exception:
        pass


def _on_photo(message):
    # PRVA linija: svaki dolazak slike vidljiv u logu zauvijek.
    print(f"[photo] ulaz chat={getattr(message.chat, 'id', '?')} "
          f"from={getattr(message.from_user, 'id', '?')} "
          f"media_group={getattr(message, 'media_group_id', None)}", flush=True)
    try:
        if not _allowed(message):
            print("[photo] odbijen — korisnik nije na whitelisti", flush=True)
            return
        try:
            file_id = message.photo[-1].file_id
            img = _download_telegram_file(file_id)
        except Exception as e:
            monitoring.error("Skidanje fotke nije uspjelo", source="racuni", exc=e)
            _bot.send_message(message.chat.id, "❌ Nisam mogao preuzeti fotografiju.")
            return
        _ingest_image(message, img, "image/jpeg")
    except Exception as e:
        _log(f"[photo] GREŠKA: {type(e).__name__}: {e}")
        monitoring.error("Greska u _on_photo", source="racuni", exc=e)
        try:
            _bot.send_message(message.chat.id,
                              "❌ Greška pri obradi fotografije. Pokušaj ponovno.")
        except Exception:
            pass


def _on_document(message):
    doc = getattr(message, "document", None)
    mime = ((doc.mime_type if doc else None) or "").lower()
    print(f"[document] ulaz chat={getattr(message.chat, 'id', '?')} "
          f"from={getattr(message.from_user, 'id', '?')} mime={mime}", flush=True)
    try:
        if not _allowed(message):
            print("[document] odbijen — korisnik nije na whitelisti", flush=True)
            return
        if not mime.startswith("image/"):
            _bot.send_message(message.chat.id,
                              "ℹ️ Pošalji dokument kao sliku (JPG/PNG) ili fotografiju "
                              "računa/primke.")
            return
        try:
            img = _download_telegram_file(doc.file_id)
        except Exception as e:
            monitoring.error("Skidanje dokumenta nije uspjelo", source="racuni", exc=e)
            _bot.send_message(message.chat.id, "❌ Nisam mogao preuzeti sliku.")
            return
        _ingest_image(message, img, mime)
    except Exception as e:
        _log(f"[document] GREŠKA: {type(e).__name__}: {e}")
        monitoring.error("Greska u _on_document", source="racuni", exc=e)
        try:
            _bot.send_message(message.chat.id,
                              "❌ Greška pri obradi dokumenta. Pokušaj ponovno.")
        except Exception:
            pass


# ==================== OBRADA (vision -> sesija -> potvrda) ====================

def _read_or_fail(chat_id, images):
    """Pozovi vision; kod greske posalji poruku i vrati None."""
    try:
        return _read_document(images)
    except json.JSONDecodeError as e:
        monitoring.warning(f"Neispravan JSON od modela: {e}", source="racuni")
        _bot.send_message(chat_id, "❌ Nisam uspio pročitati dokument (format). "
                                   "Pokušaj s jasnijom fotografijom.")
    except Exception as e:
        monitoring.error("Greska pri Claude vision pozivu", source="racuni", exc=e)
        _bot.send_message(chat_id, "❌ Greška pri čitanju dokumenta. Pokušaj ponovno.")
    return None


def _process_standalone(chat_id, user_id, who, img, mtype):
    _bot.send_message(chat_id, "🔎 Čitam dokument, trenutak...")
    data = _read_or_fail(chat_id, [(img, mtype)])
    if data is None:
        return
    try:
        total = int(data.get("broj_stranica") or 1)
    except (TypeError, ValueError):
        total = 1
    if total > 1:
        # visestranicno -> skupljaj ostale stranice
        _start_collector_for_multipage(chat_id, who, img, mtype, total)
        return
    _finish_document(chat_id, user_id, who, data, [(img, mtype)])


def _process_collected(chat_id, user_id, who, images):
    data = _read_or_fail(chat_id, images)
    if data is None:
        return
    _finish_document(chat_id, user_id, who, data, images)


def _new_session(chat_id, user_id, who, data, images):
    with _sessions_lock:
        _token_counter[0] += 1
        token = _token_counter[0]
        spec = _spec_for(data.get("vrsta"))
        sess = {
            "token": token, "chat_id": chat_id, "user_id": user_id, "who": who,
            "data": data, "images": images, "spec": spec, "vrsta": spec.vrsta,
            "gb": None, "vozac": None, "zaprimio": None,
            "stage": None, "edit_key": None,
        }
        _sessions[chat_id] = sess
    return sess


def _route_after_read(sess):
    """Nakon citanja/klasifikacije: postavi vozaca/zaprimioca i pitaj GB ili
    prikazi potvrdu."""
    driver = _safe_get_driver(sess["user_id"])
    if sess["vrsta"] == "primka":
        sess["zaprimio"] = (driver[0] if driver else sess["who"]) or ""
        sess["gb"] = None
        sess["stage"] = "need_gb"
        _bot.send_message(sess["chat_id"],
                          "Za koje vozilo (GB)? (- ako nije za konkretno vozilo)")
        return
    # racun
    if driver:
        sess["vozac"], sess["gb"] = driver
    else:
        sess["vozac"] = sess["who"]
        sess["gb"] = None
    if not sess["gb"]:
        sess["stage"] = "need_gb"
        _bot.send_message(sess["chat_id"],
                          "Koje vozilo (GB)? Napiši oznaku, npr. GB123-AB.")
        return
    sess["stage"] = "confirm"
    _send_confirm(sess)


def _finish_document(chat_id, user_id, who, data, images):
    sess = _new_session(chat_id, user_id, who, data, images)
    _route_after_read(sess)


# ==================== ODLUKE / RETRY OKO UPISA ====================

def _safe_edit(chat_id, msg_id, text, markup=None):
    try:
        _bot.edit_message_text(text, chat_id, msg_id, reply_markup=markup)
    except Exception:
        try:
            _bot.send_message(chat_id, text, reply_markup=markup)
        except Exception:
            pass


def _write_with_lock_retry(sess, chat_id, msg_id):
    total = 1 + len(_LOCK_BACKOFF)
    obavijesteno = False
    for i in range(total):
        try:
            return _write_once(sess)
        except _Locked as e:
            _log(f"upis zaključan (HTTP {e.status}); pokušaj {i + 1}/{total}")
            if not obavijesteno:
                obavijesteno = True
                _safe_edit(chat_id, msg_id,
                           "🔒 Fajl je trenutno otvoren/zaključan — "
                           "pokušavam ponovo u pozadini…")
            if i < len(_LOCK_BACKOFF):
                time.sleep(_LOCK_BACKOFF[i])
                continue
            monitoring.warning("Upis odustao (fajl zakljucan).", source="racuni")
            return False, (
                "🔒 Nisam uspio upisati — datoteka je i dalje otvorena/"
                "zaključana nakon više pokušaja.\nZatvorite je pa pošaljite ponovno.")
        except Exception as e:
            _log(f"_write_with_lock_retry neuhvacena: {type(e).__name__}: {e}")
            monitoring.error("Greska u retry petlji", source="racuni", exc=e)
            return False, "❌ Neočekivana greška pri upisu dokumenta."
    return False, "❌ Neočekivana greška pri upisu dokumenta."


def _do_ok(sess, chat_id, msg_id, force):
    """Klik ✅: DEDUPE pa upis. Radi u zasebnom daemon threadu."""
    if not force:
        try:
            if _is_duplicate(sess["spec"], sess["data"]):
                sess["busy"] = False
                _safe_edit(chat_id, msg_id, _dup_warning_text(sess),
                           markup=_dup_markup(sess["token"]))
                return
        except Exception as e:
            _log(f"dedupe provjera nije uspjela ({e}); nastavljam s upisom")
            monitoring.warning(f"Dedupe nije uspio: {e}", source="racuni")

    _do_write(sess, chat_id, msg_id)


def _do_write(sess, chat_id, msg_id):
    ok, msg = _write_with_lock_retry(sess, chat_id, msg_id)
    with _sessions_lock:
        _sessions.pop(chat_id, None)
    _safe_edit(chat_id, msg_id, msg)
    _log(f"KRAJ ({'uspjeh' if ok else 'greška'})")

    if ok and _log_note:
        try:
            data = sess["data"]
            if sess["vrsta"] == "primka":
                _log_note(chat_id, f"Primka upisana: {data.get('firma') or ''} "
                          f"{_fmt_num(_parse_num(data.get('ukupno_eur')))} EUR")
            else:
                _log_note(chat_id, f"Račun upisan: {data.get('izdavatelj') or ''} "
                          f"{_fmt_num(_parse_num(data.get('ukupno_eur')))} EUR")
        except Exception:
            pass


def _reread_as(sess, chat_id, msg_id, other_vrsta):
    """🔄 Promijeni vrstu: ponovno procitaj iste slike kao drugu vrstu."""
    try:
        data = _read_document(sess["images"], force_vrsta=other_vrsta)
    except Exception as e:
        monitoring.error("Ponovno citanje (promjena vrste) nije uspjelo",
                         source="racuni", exc=e)
        _safe_edit(chat_id, msg_id, "❌ Greška pri ponovnom čitanju. Pokušaj opet.")
        with _sessions_lock:
            _sessions.pop(chat_id, None)
        return
    sess["data"] = data
    sess["spec"] = _spec_for(other_vrsta)
    sess["vrsta"] = sess["spec"].vrsta
    sess["busy"] = False
    # GB/vozac/zaprimioca ponovno posloziti prema novoj vrsti
    try:
        _bot.delete_message(chat_id, msg_id)
    except Exception:
        pass
    _route_after_read(sess)


# ==================== CALLBACKS ====================

def _on_callback(c):
    if c.from_user.id not in _allowed_users:
        _bot.answer_callback_query(c.id)
        return
    try:
        _bot.answer_callback_query(c.id)
    except Exception:
        pass

    try:
        _, action, tok = c.data.split("_")
        token = int(tok)
    except ValueError:
        return

    chat_id = c.message.chat.id
    with _sessions_lock:
        sess = _sessions.get(chat_id)
        if not sess or sess["token"] != token:
            sess = None

    if not sess:
        try:
            _bot.edit_message_text("Ovaj dokument više nije aktivan.",
                                   chat_id, c.message.message_id)
        except Exception:
            pass
        return

    if sess.get("busy"):
        return

    if action == "no":
        with _sessions_lock:
            _sessions.pop(chat_id, None)
        try:
            _bot.edit_message_text("❌ Dokument odbačen.", chat_id, c.message.message_id)
        except Exception:
            pass
        return

    if action == "ed":
        sess["stage"] = "edit_which"
        polja = ", ".join(_edit_field_names(sess["vrsta"]))
        _bot.send_message(
            chat_id,
            "Koje polje želiš ispraviti? Napiši naziv, npr. 'oib'.\n"
            "(Pojedine stavke se ne uređuju — ako su krive, odbaci i pošalji "
            "jasniju fotku.)\n\n"
            f"Moguća polja: {polja}")
        return

    if action == "ty":
        other = "racun" if sess["vrsta"] == "primka" else "primka"
        sess["busy"] = True
        msg_id = c.message.message_id
        _safe_edit(chat_id, msg_id,
                   f"🔄 Ponovno čitam kao {_spec_for(other).naziv}…")
        threading.Thread(target=_reread_as,
                         args=(sess, chat_id, msg_id, other), daemon=True).start()
        return

    if action in ("ok", "fo"):
        force = (action == "fo")
        msg_id = c.message.message_id
        sess["busy"] = True
        _safe_edit(chat_id, msg_id,
                   "⏳ Upisujem…" if force else "⏳ Provjeravam duplikat i upisujem…")
        threading.Thread(target=_do_ok, args=(sess, chat_id, msg_id, force),
                         daemon=True).start()
        return


def _on_collect_cb(c):
    """Gumb 'Obradi dokument' (doc_go)."""
    if c.from_user.id not in _allowed_users:
        _bot.answer_callback_query(c.id)
        return
    try:
        _bot.answer_callback_query(c.id)
    except Exception:
        pass
    _trigger_process(c.message.chat.id, c.from_user.id, c.from_user.full_name,
                     edit_msg_id=c.message.message_id)


def _trigger_process(chat_id, user_id, who, edit_msg_id=None):
    with _collectors_lock:
        col = _collectors.pop(chat_id, None)
    if not col or not col.get("images"):
        if edit_msg_id:
            _safe_edit(chat_id, edit_msg_id, "Nema stranica za obradu.")
        else:
            _bot.send_message(chat_id, "Nema dokumenta u obradi.")
        return
    n = len(col["images"])
    who = col.get("who") or who
    if edit_msg_id:
        _safe_edit(chat_id, edit_msg_id, f"🔎 Čitam {n} stranica, trenutak...")
    else:
        _bot.send_message(chat_id, f"🔎 Čitam {n} stranica, trenutak...")
    threading.Thread(target=_process_collected,
                     args=(chat_id, user_id, who, col["images"]), daemon=True).start()


# ==================== EDIT POLJA ====================

_RACUN_EDIT_ALIASES = {
    "datum": ("data", "datum"), "vrijeme": ("data", "vrijeme"),
    "izdavatelj": ("data", "izdavatelj"), "oib": ("data", "oib"),
    "broj racuna": ("data", "broj_racuna"), "broj_racuna": ("data", "broj_racuna"),
    "broj": ("data", "broj_racuna"), "lokacija": ("data", "lokacija"),
    "ukupno": ("data", "ukupno_eur"), "ukupno_eur": ("data", "ukupno_eur"),
    "iznos": ("data", "ukupno_eur"), "pdv": ("data", "pdv_iznos"),
    "vozac": ("sess", "vozac"), "gb": ("sess", "gb"), "vozilo": ("sess", "gb"),
}

_PRIMKA_EDIT_ALIASES = {
    "datum": ("data", "datum_dokumenta"),
    "datum dokumenta": ("data", "datum_dokumenta"),
    "firma": ("data", "firma"), "oib": ("data", "oib"),
    "br racuna": ("data", "br_racuna"), "br_racuna": ("data", "br_racuna"),
    "broj": ("data", "br_racuna"), "dospijece": ("data", "dospijece"),
    "dospijeće": ("data", "dospijece"),
    "ukupno": ("data", "ukupno_eur"), "ukupno_eur": ("data", "ukupno_eur"),
    "iznos": ("data", "ukupno_eur"),
    "gb": ("sess", "gb"), "vozilo": ("sess", "gb"),
    "zaprimio": ("sess", "zaprimio"),
}

_NUM_FIELDS = {"ukupno_eur", "pdv_iznos"}


def _edit_aliases(vrsta):
    return _PRIMKA_EDIT_ALIASES if vrsta == "primka" else _RACUN_EDIT_ALIASES


def _edit_field_names(vrsta):
    if vrsta == "primka":
        return ["datum", "firma", "oib", "br racuna", "dospijece", "ukupno",
                "gb", "zaprimio"]
    return ["datum", "vrijeme", "izdavatelj", "oib", "broj racuna", "lokacija",
            "ukupno", "pdv", "vozac", "gb"]


# ==================== PENDING TEXT (zove main.handle) ====================

def handle_text(message):
    """Vraca True ako je poruka obradjena ovdje (pending/kolektor/komanda)."""
    if not _allowed(message):
        return False

    text = (message.text or "").strip()
    low = text.lower()

    if low.startswith("/vozac_dodaj"):
        _cmd_vozac_dodaj(message)
        return True
    if low.startswith("/vozac_lista"):
        _cmd_vozac_lista(message)
        return True

    # /gotovo -> obradi skupljene stranice (ako ih ima)
    if low.startswith("/gotovo"):
        with _collectors_lock:
            has = message.chat.id in _collectors
        if has:
            _trigger_process(message.chat.id, message.from_user.id,
                             message.from_user.full_name)
            return True
        return False  # nema kolektora -> pusti dalje

    with _sessions_lock:
        sess = _sessions.get(message.chat.id)
    if not sess:
        return False

    stage = sess.get("stage")
    if stage == "need_gb":
        sess["gb"] = text
        sess["stage"] = "confirm"
        _send_confirm(sess)
        return True

    if stage == "edit_which":
        key = _edit_aliases(sess["vrsta"]).get(low)
        if not key:
            _bot.send_message(message.chat.id,
                              "Ne prepoznajem to polje. Pokušaj npr. 'oib' ili 'ukupno'.")
            return True
        sess["edit_key"] = key
        sess["stage"] = "edit_value"
        _bot.send_message(message.chat.id, f"Nova vrijednost za '{low}':")
        return True

    if stage == "edit_value":
        target, field = sess["edit_key"]
        if target == "sess":
            sess[field] = text
        else:
            if field in _NUM_FIELDS:
                num = _parse_num(text)
                sess["data"][field] = num if num is not None else text
            else:
                sess["data"][field] = text
        sess["edit_key"] = None
        sess["stage"] = "confirm"
        _bot.send_message(message.chat.id, "✅ Ažurirano.")
        _send_confirm(sess)
        return True

    return False


def _cmd_vozac_dodaj(message):
    # /vozac_dodaj <telegram_id> <GB> <ime i prezime>
    parts = message.text.split(maxsplit=3)
    if len(parts) < 4:
        _bot.reply_to(message,
                      "Format: /vozac_dodaj <telegram_id> <GB> <ime i prezime>\n"
                      "Npr: /vozac_dodaj 5191857104 GB123-AB Ivan Horvat")
        return
    try:
        uid = int(parts[1])
    except ValueError:
        _bot.reply_to(message, "Telegram ID mora biti broj.")
        return
    gb = parts[2]
    ime = parts[3].strip()
    try:
        upsert_driver(uid, ime, gb)
    except Exception as e:
        monitoring.error("Upsert vozaca nije uspio", source="racuni", exc=e)
        _bot.reply_to(message, "❌ Greška pri spremanju vozača.")
        return
    _bot.reply_to(message, f"✅ Vozač spremljen:\n{ime} — {gb} (ID {uid})")


def _cmd_vozac_lista(message):
    try:
        rows = list_drivers()
    except Exception as e:
        monitoring.error("Lista vozaca nije uspjela", source="racuni", exc=e)
        _bot.reply_to(message, "❌ Greška pri dohvatu vozača.")
        return
    if not rows:
        _bot.reply_to(message, "Nema spremljenih vozača. Dodaj s /vozac_dodaj.")
        return
    lines = ["🚚 Vozači:"]
    for r in rows:
        lines.append(f"• {r['ime_vozaca']} — {r['gb_vozila'] or '?'} "
                     f"(ID {r['telegram_user_id']})")
    _bot.reply_to(message, "\n".join(lines))


# ==================== REGISTRACIJA HANDLERA ====================

def register(bot):
    """Registrira photo/document/callback handlere. Text (pending stanja,
    /gotovo, /vozac_*) ide preko main.handle -> handle_text()."""
    bot.message_handler(content_types=["photo"])(_on_photo)
    bot.message_handler(content_types=["document"])(_on_document)
    bot.callback_query_handler(func=lambda c: c.data.startswith("rc_"))(_on_callback)
    bot.callback_query_handler(func=lambda c: c.data.startswith("doc_"))(_on_collect_cb)
