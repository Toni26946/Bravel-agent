# ============================================================
#  whatsapp_meni.py — WhatsApp kao UPRAVLJAČKA PLOČA za vozače/radnike
#
#  Ovlašteni radnik (WHATSAPP_ALLOWED) piše poslovnom broju → dobije IZBORNIK
#  (interaktivna lista) s opcijama koje sam odrađuje:
#    🧾 Račun/primka   → postojeći tok (whatsapp_racuni)
#    📍 Gdje je vozilo → Mobilisis pozicija (ovisi o IP whitelistu)
#    🛠️ Prijava kvara  → prosljeđuje se vlasnicima na Telegram
#    🕒 Evidencija sati → dolazak/odlazak (tablica wa_sati)
#    🔔 Moj podsjetnik → radnik si sam postavi podsjetnik (tablica wa_podsjetnici)
#    ℹ️ Pomoć
#
#  Ovaj modul je NOVI ULAZ za dolazne poruke ovlaštenih (main.wa_dolazna_poruka):
#  odlučuje ide li poruka u tok računa, u meni-tok, ili prikazuje izbornik.
#  Ovisnosti (Mobilisis lookup, obavijest vlasnicima) ubrizgava main preko setup().
# ============================================================

import re
import threading
from datetime import timedelta

import racuni
import whatsapp
import whatsapp_racuni
import graph_client
import monitoring

# Evidencija sati → SharePoint Excel (uz bazu wa_sati)
_SATI_FILE = "Evidencija_sati.xlsx"
_SATI_TABLE = "Sati"
_SATI_COLS = ["Datum", "Vrijeme", "Ime", "Broj", "Tip"]

_stanje = {}            # broj → {"tok": ..., ...}  (meni-tokovi, ne računi)
_lock = threading.RLock()

_gdje_lookup = None     # fn(query)->str  (Mobilisis pozicija; ubrizgava main)
_obavijesti = None      # fn(text)->None  (šalje vlasnicima na Telegram)

_ROWS = [
    ("meni_racun",      "🧾 Račun / primka",  "Pošalji fotografiju dokumenta"),
    ("meni_lokacija",   "📍 Gdje je vozilo",  "Trenutna pozicija kamiona"),
    ("meni_kvar",       "🛠️ Prijava kvara",   "Javi kvar ili problem"),
    ("meni_sati",       "🕒 Evidencija sati", "Dolazak / odlazak"),
    ("meni_podsjetnik", "🔔 Moj podsjetnik",  "Postavi si podsjetnik"),
    ("meni_pomoc",      "ℹ️ Pomoć",           "Kako koristiti"),
]

_MENI_RIJECI = {"meni", "menu", "izbornik", "start", "bok", "pozdrav", "hej",
                "hi", "hello", "?"}

# Utipkane ključne riječi → ista akcija kao klik na stavku izbornika.
_KEYWORD_RID = {
    "racun": "meni_racun", "račun": "meni_racun", "primka": "meni_racun",
    "primku": "meni_racun", "racuni": "meni_racun", "računi": "meni_racun",
    "lokacija": "meni_lokacija", "gdje": "meni_lokacija", "vozilo": "meni_lokacija",
    "kvar": "meni_kvar", "kvara": "meni_kvar", "problem": "meni_kvar",
    "sati": "meni_sati", "evidencija": "meni_sati", "dolazak": "meni_sati",
    "odlazak": "meni_sati",
    "podsjetnik": "meni_podsjetnik", "podsjetnici": "meni_podsjetnik",
    "pomoc": "meni_pomoc", "pomoć": "meni_pomoc", "help": "meni_pomoc",
}

_POMOC = (
    "ℹ️ *Bravel — upute za korištenje*\n\n"
    "Napiši „meni” bilo kad da otvoriš izbornik. Opcije:\n\n"
    "🧾 *Račun / primka*\n"
    "Slikaj dokument i pošalji fotografiju. Pitam te za vozilo (GB) i pokažem što "
    "sam pročitao — potvrdiš ✅. Više stranica? Šalji sve pa napiši „gotovo”. "
    "Krivo prepoznata vrsta? Napiši „vrsta”. Krivo pročitano? Napiši „ispravi”.\n\n"
    "📍 *Gdje je vozilo*\n"
    "Upiši GB ili registraciju i dobiješ trenutnu lokaciju.\n\n"
    "🛠️ *Prijava kvara*\n"
    "Opiši kvar ili problem — javit ćemo se.\n\n"
    "🕒 *Evidencija sati*\n"
    "🟢 Dolazak kad počneš, 🔴 Odlazak kad završiš.\n\n"
    "🔔 *Moj podsjetnik*\n"
    "Upiši kada (npr. „sutra 08:00”) pa tekst — javim ti u to vrijeme.\n\n"
    "💡 Savjet: pritisni i zadrži ovu poruku pa odaberi *Prikvači* (pin) — tako ti "
    "upute uvijek stoje na vrhu razgovora.")


def setup(gdje_lookup=None, obavijesti=None):
    """main ubrizgava: gdje_lookup(query)->str (Mobilisis), obavijesti(text)->None
    (šalje vlasnicima na Telegram). Kreira i potrebne tablice."""
    global _gdje_lookup, _obavijesti
    _gdje_lookup = gdje_lookup
    _obavijesti = obavijesti
    _ensure_tables()


def _ensure_tables():
    try:
        with racuni._db() as conn:
            conn.execute("CREATE TABLE IF NOT EXISTS wa_sati ("
                         "id INTEGER PRIMARY KEY AUTOINCREMENT, broj TEXT, ime TEXT, "
                         "tip TEXT, ts INTEGER)")
            conn.execute("CREATE TABLE IF NOT EXISTS wa_podsjetnici ("
                         "id INTEGER PRIMARY KEY AUTOINCREMENT, broj TEXT, ime TEXT, "
                         "tekst TEXT, ts INTEGER, poslan INTEGER DEFAULT 0)")
    except Exception as e:
        monitoring.warning(f"WA meni: init tablica nije uspio: {e}", source="wa_meni")


def _set(frm, tok, **kw):
    with _lock:
        _stanje[frm] = {"tok": tok, **kw}


def _clear(frm):
    with _lock:
        _stanje.pop(frm, None)


# ==================== ULAZ ====================

def obradi(frm, ime, msg):
    """Glavni ulaz za dolaznu poruku ovlaštenog radnika."""
    try:
        _obradi(frm, ime, msg)
    except Exception as e:
        monitoring.error("WhatsApp meni: greška u obradi", source="wa_meni", exc=e)
        try:
            whatsapp.send_text(frm, "❌ Došlo je do greške. Napiši „meni” za izbornik.")
        except Exception:
            pass


def _obradi(frm, ime, msg):
    tip = msg.get("type")

    # 1) Ako je aktivan tok RAČUNA (sesija/sakupljanje/obrada) → SVE njemu,
    #    UKLJUČUJUĆI gumbe potvrde (Upiši/Ispravi/Odbaci). Mora biti prije
    #    obrade izbornika, inače potvrda računa završi kao „nepoznata” stavka
    #    izbornika (klik „Upiši” ne bi upisao).
    if whatsapp_racuni.zauzet(frm):
        whatsapp_racuni.handle(frm, ime, msg)
        return

    # 2) Interaktivni odgovor (izbornik/gumbi)
    if tip == "interactive":
        inter = msg.get("interactive") or {}
        rid = ((inter.get("list_reply") or {}).get("id")
               or (inter.get("button_reply") or {}).get("id") or "")
        _izbor(frm, ime, rid)
        return

    # 3) Ako smo u nekom meni-toku (lokacija/kvar/podsjetnik) → nastavi ga
    with _lock:
        st = _stanje.get(frm)
    if st and st.get("tok"):
        _tok(frm, ime, msg, st)
        return

    # 4) Slika/dokument bez aktivnog toka → računi (najčešći slučaj)
    if tip in ("image", "document"):
        whatsapp_racuni.handle(frm, ime, msg)
        return

    # 5) Tekst: utipkana ključna riječ pokreće tok, inače prikaži izbornik.
    if tip == "text":
        low = (msg.get("text") or {}).get("body", "").strip().lower()
        rid = _KEYWORD_RID.get(low)
        if rid:
            _izbor(frm, ime, rid)
            return
    _posalji_meni(frm, ime)


def _posalji_meni(frm, ime):
    poz = f"Bok {ime}! " if ime and not str(ime).isdigit() else ""
    res = whatsapp.send_list(frm, f"{poz}Što trebaš? Odaberi opciju:",
                             "Izbornik", _ROWS, header="Bravel")
    if not res.get("ok"):
        # Ako lista ne prođe (npr. izvan 24 h), barem tekstualni izbornik.
        whatsapp.send_text(frm,
            "Bravel — što trebaš? Napiši:\n"
            "• „racun” (pošalji fotografiju)\n• „lokacija”\n• „kvar”\n"
            "• „sati”\n• „podsjetnik”")


# ==================== ODABIR IZ IZBORNIKA ====================

def _izbor(frm, ime, rid):
    if rid in ("meni_racun", "racun"):
        _clear(frm)
        whatsapp.send_text(frm, "🧾 Pošalji fotografiju računa ili primke. Za više "
                                "stranica šalji redom pa napiši „gotovo”.")
        return
    if rid in ("meni_lokacija", "lokacija"):
        _set(frm, "lokacija")
        whatsapp.send_text(frm, "📍 Koje vozilo? Napiši GB ili registraciju "
                                "(npr. GB123 ili ZG1234AB).")
        return
    if rid in ("meni_kvar", "kvar"):
        _set(frm, "kvar")
        whatsapp.send_text(frm, "🛠️ Opiši kvar ili problem (u jednoj poruci; može "
                                "i fotografija).")
        return
    if rid in ("meni_sati", "sati"):
        _clear(frm)
        whatsapp.send_buttons(frm, "🕒 Evidencija radnog vremena:",
                              [("sati_dolazak", "🟢 Dolazak"),
                               ("sati_odlazak", "🔴 Odlazak")])
        return
    if rid in ("meni_podsjetnik", "podsjetnik"):
        _set(frm, "pods_kada")
        whatsapp.send_text(frm, "🔔 Kada da te podsjetim? Npr:\n"
                                "• sutra 08:00\n• 22.07 14:30\n• 18:00 (danas)")
        return
    if rid in ("meni_pomoc", "pomoc", "pomoć", "help"):
        _clear(frm)
        whatsapp.send_text(frm, _POMOC)
        return
    if rid == "sati_dolazak":
        _zabiljezi_sat(frm, ime, "dolazak")
        return
    if rid == "sati_odlazak":
        _zabiljezi_sat(frm, ime, "odlazak")
        return
    # nepoznat id → izbornik
    _posalji_meni(frm, ime)


# ==================== TOKOVI ====================

def _tok(frm, ime, msg, st):
    tok = st.get("tok")
    tip = msg.get("type")
    tekst = (msg.get("text") or {}).get("body", "").strip() if tip == "text" else ""
    low = tekst.lower()

    # „meni” u bilo kojem toku → izlaz na izbornik
    if low in _MENI_RIJECI:
        _clear(frm)
        _posalji_meni(frm, ime)
        return

    if tok == "lokacija":
        if tip != "text":
            whatsapp.send_text(frm, "Napiši GB ili registraciju vozila (tekstom).")
            return
        _clear(frm)
        whatsapp.send_text(frm, "🔎 Tražim lokaciju…")
        try:
            txt = _gdje_lookup(tekst) if _gdje_lookup else None
        except Exception as e:
            monitoring.warning(f"WA lokacija lookup: {e}", source="wa_meni")
            txt = None
        whatsapp.send_text(frm, txt or "📍 Lokacija trenutno nedostupna "
                                       "(sustav praćenja privremeno nedostupan).")
        return

    if tok == "kvar":
        if tip == "text":
            opis = tekst or "(bez opisa)"
        elif tip in ("image", "document"):
            opis = "(radnik je poslao fotografiju kvara)"
        else:
            opis = f"(poruka tipa {tip})"
        _clear(frm)
        poruka = (f"🛠️ PRIJAVA KVARA (WhatsApp)\n"
                  f"Od: {ime or frm} ({frm})\n\n{opis}")
        if _obavijesti:
            try:
                _obavijesti(poruka)
            except Exception as e:
                monitoring.warning(f"WA kvar obavijest: {e}", source="wa_meni")
        whatsapp.send_text(frm, "✅ Prijava poslana vlasnicima. Javit ćemo se. Hvala!")
        return

    if tok == "pods_kada":
        if tip != "text":
            whatsapp.send_text(frm, "Napiši vrijeme, npr. „sutra 08:00”.")
            return
        dt = _parse_kada(tekst)
        if not dt:
            whatsapp.send_text(frm, "Ne razumijem vrijeme. Primjeri:\n"
                                    "• sutra 08:00\n• 22.07 14:30\n• 18:00")
            return
        _set(frm, "pods_sto", ts=int(dt.timestamp()),
             kada_txt=dt.strftime("%d.%m. u %H:%M"))
        whatsapp.send_text(frm, f"🔔 Za {dt.strftime('%d.%m. u %H:%M')}. "
                                "Što da ti napišem kao podsjetnik?")
        return

    if tok == "pods_sto":
        if tip != "text":
            whatsapp.send_text(frm, "Napiši tekst podsjetnika.")
            return
        ts = st.get("ts")
        kada = st.get("kada_txt", "")
        _clear(frm)
        try:
            with racuni._db() as conn:
                conn.execute("INSERT INTO wa_podsjetnici (broj, ime, tekst, ts, poslan) "
                             "VALUES (?, ?, ?, ?, 0)", (frm, ime or frm, tekst, ts))
            whatsapp.send_text(frm, f"✅ Podsjetnik postavljen za {kada}:\n„{tekst}”")
        except Exception as e:
            monitoring.error("WA podsjetnik upis", source="wa_meni", exc=e)
            whatsapp.send_text(frm, "❌ Nisam uspio spremiti podsjetnik.")
        return

    # nepoznat tok
    _clear(frm)
    _posalji_meni(frm, ime)


def _sati_workbook_bytes(rows):
    """Novi xlsx s tablicom 'Sati' (za prvi upis kad datoteka ne postoji)."""
    import io
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = _SATI_TABLE
    ws.append(_SATI_COLS)
    for r in rows:
        ws.append(r)
    last = get_column_letter(len(_SATI_COLS))
    tab = Table(displayName=_SATI_TABLE, ref=f"A1:{last}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sati_u_excel(now, ime, broj, tip):
    """Dopiši redak u Evidencija_sati.xlsx na SharePointu (best-effort).
    Kreira datoteku ako ne postoji. Zove se u pozadinskom threadu."""
    if not graph_client.is_configured():
        return
    row = [now.strftime("%d.%m.%Y"), now.strftime("%H:%M"),
           ime, broj, "Dolazak" if tip == "dolazak" else "Odlazak"]
    try:
        if not graph_client.file_exists(_SATI_FILE):
            graph_client.upload_file(_SATI_FILE, _sati_workbook_bytes([row]))
            return
        graph_client.append_or_fill_table_rows(_SATI_FILE, _SATI_TABLE, [row])
    except Exception as e:
        monitoring.warning(f"WA sati → SharePoint nije uspio: {e}", source="wa_meni")


def _zabiljezi_sat(frm, ime, tip):
    _clear(frm)
    now = racuni._now()
    try:
        with racuni._db() as conn:
            conn.execute("INSERT INTO wa_sati (broj, ime, tip, ts) VALUES (?, ?, ?, ?)",
                         (frm, ime or frm, tip, int(now.timestamp())))
    except Exception as e:
        monitoring.error("WA sati upis", source="wa_meni", exc=e)
        whatsapp.send_text(frm, "❌ Nisam uspio zabilježiti.")
        return
    rijec = "Dolazak" if tip == "dolazak" else "Odlazak"
    kada = now.strftime("%d.%m.%Y %H:%M")
    whatsapp.send_text(frm, f"✅ {rijec} zabilježen: {kada}.")
    # SharePoint Excel upis u pozadini (ne blokira odgovor radniku).
    threading.Thread(target=_sati_u_excel, args=(now, ime or frm, frm, tip),
                     daemon=True).start()
    if _obavijesti:
        try:
            _obavijesti(f"🕒 {rijec} — {ime or frm} ({frm}) u {kada}")
        except Exception:
            pass


# ==================== PARSIRANJE VREMENA ====================

def _parse_kada(s):
    """Vrati tz-aware datetime za podsjetnik ili None. Podržava:
    'sutra HH:MM', 'danas HH:MM', 'DD.MM[.YYYY] HH:MM', 'HH:MM'."""
    now = racuni._now()
    low = s.strip().lower()

    def mk(day, h, mi):
        if h > 23 or mi > 59:
            return None
        try:
            return day.replace(hour=h, minute=mi, second=0, microsecond=0)
        except Exception:
            return None

    m = re.search(r"\bsutra\b\D*(\d{1,2})[:.h](\d{2})", low)
    if m:
        return mk(now + timedelta(days=1), int(m.group(1)), int(m.group(2)))

    m = re.search(r"\bdanas\b\D*(\d{1,2})[:.h](\d{2})", low)
    if m:
        return mk(now, int(m.group(1)), int(m.group(2)))

    m = re.search(r"(\d{1,2})\.(\d{1,2})\.?(\d{4})?\D+(\d{1,2})[:.h](\d{2})", low)
    if m:
        dd, mm, yy, h, mi = m.groups()
        try:
            dt = now.replace(year=int(yy) if yy else now.year, month=int(mm),
                             day=int(dd), hour=int(h), minute=int(mi),
                             second=0, microsecond=0)
        except Exception:
            return None
        if not yy and dt < now:
            try:
                dt = dt.replace(year=dt.year + 1)
            except Exception:
                pass
        return dt

    m = re.search(r"\b(\d{1,2})[:.h](\d{2})\b", low)
    if m:
        dt = mk(now, int(m.group(1)), int(m.group(2)))
        if dt and dt <= now:
            dt = dt + timedelta(days=1)
        return dt

    return None


# ==================== ISPORUKA PODSJETNIKA (zove scheduler) ====================

def posalji_dospjele():
    """Pošalji dospjele radničke podsjetnike (zove se iz check_reminders petlje).
    Unutar 24 h prozora ide send_text; izvan → predložak podsjetnik_opci."""
    now_ts = int(racuni._now().timestamp())
    try:
        with racuni._db() as conn:
            due = conn.execute("SELECT id, broj, tekst FROM wa_podsjetnici "
                               "WHERE poslan = 0 AND ts <= ?", (now_ts,)).fetchall()
    except Exception as e:
        monitoring.warning(f"WA podsjetnici dohvat: {e}", source="wa_meni")
        return

    for row in due:
        rid, broj, tekst = row[0], row[1], row[2]
        # Označi poslan PRIJE slanja (da se ne ponavlja na grešci).
        try:
            with racuni._db() as conn:
                conn.execute("UPDATE wa_podsjetnici SET poslan = 1 WHERE id = ?", (rid,))
        except Exception:
            pass
        res = whatsapp.send_text(broj, f"🔔 Podsjetnik: {tekst}")
        if not res.get("ok"):
            # vjerojatno izvan 24 h prozora → probaj odobreni predložak
            r2 = whatsapp.send_template(
                broj, "podsjetnik_opci", "hr",
                components=[{"type": "body",
                             "parameters": [{"type": "text", "text": tekst}]}])
            if not r2.get("ok"):
                monitoring.warning(
                    f"WA podsjetnik nije isporučen ({broj}): {whatsapp.opisi_gresku(r2)}",
                    source="wa_meni")
