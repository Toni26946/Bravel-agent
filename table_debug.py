# ============================================================
#  TABLE_DEBUG - dijagnostika stvarnog stanja Racuni_terena.xlsx
#
#  Ispisuje STVARNO stanje tablice "Racuni" onako kako ga vidi
#  Microsoft Graph (workbook API) I openpyxl (isti skinuti fajl),
#  da vidimo je li prazan red UNUTAR tablice (duh koji append
#  postuje) ili je tablica uredna a append krivo cilja.
#
#  Pokretanje na Fly:
#     fly ssh console
#     python table_debug.py
#  (Graph secrets su vec u okolini stroja.)
# ============================================================

import io

import graph_client as gc

FILENAME = "Racuni_terena.xlsx"
TABLE = "Racuni"


def hr(title):
    print("\n" + "=" * 60)
    print(title)
    print("=" * 60)


def show_rows(values, label="redak"):
    """Ispisi svaki redak 2D liste s oznakom je li prazan."""
    if not values:
        print("  (nema redaka)")
        return
    for i, row in enumerate(values):
        empty = all(c is None or str(c).strip() == "" for c in row)
        flag = "PRAZAN" if empty else "  data"
        print(f"  [{i}] {flag} | {row!r}")


def graph_view():
    hr("1) GRAPH WORKBOOK API")
    if not gc.is_configured():
        print("  !! Graph NIJE konfiguriran (nedostaju kredencijali/msal).")
        return

    gc.ensure_token()
    item_id = gc.get_item_id(FILENAME)
    drive_id = gc.get_drive_id()
    base = f"{gc.GRAPH_ROOT}/drives/{drive_id}/items/{item_id}/workbook"
    print(f"  drive_id = {drive_id}")
    print(f"  item_id  = {item_id}")

    # -- tablica: raspon (ref) --
    try:
        r = gc._request("GET", f"{base}/tables/{TABLE}/range").json()
        print(f"\n  tables/{TABLE}/range:")
        print(f"    address     = {r.get('address')}")
        print(f"    rowCount    = {r.get('rowCount')}")
        print(f"    columnCount = {r.get('columnCount')}")
    except Exception as e:
        print(f"  !! range greska: {e}")

    # -- header --
    try:
        h = gc._request("GET", f"{base}/tables/{TABLE}/headerRowRange").json()
        print(f"\n  tables/{TABLE}/headerRowRange:")
        print(f"    address = {h.get('address')}")
        print(f"    values  = {h.get('values')}")
    except Exception as e:
        print(f"  !! headerRowRange greska: {e}")

    # -- dataBodyRange: ovo append gleda --
    try:
        d = gc._request("GET", f"{base}/tables/{TABLE}/dataBodyRange").json()
        print(f"\n  tables/{TABLE}/dataBodyRange:")
        print(f"    address  = {d.get('address')}")
        print(f"    rowCount = {d.get('rowCount')}")
        print("    values (svaki redak, ukljucivo prazne):")
        show_rows(d.get("values") or [])
    except Exception as e:
        print(f"  !! dataBodyRange greska: {e}")

    # -- rows kolekcija: sto tablica prijavljuje kao retke --
    try:
        rc = gc._request("GET", f"{base}/tables/{TABLE}/rows").json()
        items = rc.get("value") or []
        print(f"\n  tables/{TABLE}/rows: prijavljeno {len(items)} redaka")
        for it in items:
            vals = it.get("values") or [[]]
            row = vals[0] if vals else []
            empty = all(c is None or str(c).strip() == "" for c in row)
            flag = "PRAZAN" if empty else "  data"
            print(f"    index={it.get('index')} {flag} | {row!r}")
    except Exception as e:
        print(f"  !! rows greska: {e}")

    # -- usedRange lista --
    try:
        w = gc._request("GET", f"{base}/tables/{TABLE}/worksheet").json()
        ws_id = w.get("id")
        print(f"\n  worksheet: name={w.get('name')!r} id={ws_id}")
        seg = gc.requests.utils.quote(ws_id, safe="")
        u = gc._request("GET", f"{base}/worksheets/{seg}/usedRange").json()
        print(f"  worksheets/usedRange:")
        print(f"    address  = {u.get('address')}")
        print(f"    rowCount = {u.get('rowCount')}")
        print("    values (svaki redak):")
        show_rows(u.get("values") or [])
    except Exception as e:
        print(f"  !! usedRange greska: {e}")


def openpyxl_view():
    hr("2) OPENPYXL (isti skinuti fajl)")
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"  !! openpyxl nedostupan: {e}")
        return
    try:
        content = gc.download_file(FILENAME)
    except Exception as e:
        print(f"  !! download greska: {e}")
        return

    wb = load_workbook(io.BytesIO(content))
    ws = wb["Racuni"] if "Racuni" in wb.sheetnames else wb.active
    print(f"  sheet            = {ws.title!r}")
    print(f"  ws.dimensions    = {ws.dimensions}")
    print(f"  ws.max_row       = {ws.max_row}")
    print(f"  ws.max_column    = {ws.max_column}")
    print(f"  sheetnames       = {wb.sheetnames}")

    tabs = getattr(ws, "tables", {})
    print(f"  tablice na listu = {list(tabs.keys())}")
    for name, t in tabs.items():
        # t moze biti Table objekt ili string ref, ovisno o verziji
        ref = getattr(t, "ref", t)
        print(f"    tablica {name!r} ref = {ref}")

    print("\n  SVI redci 1..max_row (ukljucivo prazne):")
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        empty = all(c is None or str(c).strip() == "" for c in row)
        flag = "PRAZAN" if empty else "  data"
        print(f"    row {r} {flag} | {row!r}")


if __name__ == "__main__":
    print(f"DIJAGNOSTIKA: {FILENAME}  (tablica {TABLE!r})")
    graph_view()
    openpyxl_view()
    hr("KRAJ")
